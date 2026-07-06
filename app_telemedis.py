import streamlit as st
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Flask, request, jsonify
import threading
import time
import random
import requests

# =====================================================================
# 🌟 KECERDASAN BUATAN CLOUD: GOOGLE GEMINI AI
# =====================================================================
import google.generativeai as genai

# ---> MASUKKAN KUNCI API GOOGLE GEMINI ANDA DI DALAM TANDA KUTIP INI:
GEMINI_API_KEY = "AQ.Ab8RN6Jfmz2aRMHoeNJRJ8JA1LL9HSIqXw0Rmcio1MWTF9sn9Q"

if GEMINI_API_KEY != "MASUKKAN_API_KEY_GEMINI_DI_SINI":
    genai.configure(api_key=GEMINI_API_KEY)

def minta_diagnosis_ai_gemini(suhu, hr, spo2, riwayat, nama):
    if GEMINI_API_KEY == "MASUKKAN_API_KEY_GEMINI_DI_SINI":
        return "⚠️ Kunci API Google Gemini belum dimasukkan."
    try:
        # GANTI BAGIAN INI dengan salah satu model dari daftar yang muncul di layar Anda
        model_pilihan = 'gemini-2.5-flash' 
        
        model = genai.GenerativeModel(model_pilihan)
        
        prompt = f"Analisis medis untuk {nama}. Riwayat: {riwayat}. Data: Suhu {suhu}C, HR {hr}BPM, SpO2 {spo2}%. Berikan diagnosis singkat."
        
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        return f"Terjadi kesalahan teknis: {e}"
# =====================================================================
# KECERDASAN BUATAN LOKAL: AI PREDIKSI TREN
# =====================================================================
def prediksi_ai_tren(df_medis):
    if len(df_medis) < 5:
        return "⏳ AI Lokal sedang mengumpulkan data klinis untuk menganalisis tren..."
    
    tren_data = df_medis.head(5).iloc[::-1]
    spo2_awal = tren_data['spo2'].iloc[0]
    spo2_akhir = tren_data['spo2'].iloc[-1]
    hr_awal = tren_data['detak_jantung'].iloc[0]
    hr_akhir = tren_data['detak_jantung'].iloc[-1]
    
    perubahan_spo2 = spo2_akhir - spo2_awal
    perubahan_hr = hr_akhir - hr_awal
    
    if perubahan_spo2 <= -3 and spo2_akhir <= 94:
        return "🚨 AI LOKAL: Waspada! Terdeteksi tren penurunan oksigen secara agresif. Risiko hipoksia mendadak!"
    elif perubahan_hr >= 12 and hr_akhir > 95:
        return "⚠️ AI LOKAL: Detak jantung menunjukkan tren lonjakan konsisten. Indikasi syok atau demam tinggi."
    elif spo2_akhir >= 96 and 60 <= hr_akhir <= 90:
        return "✅ AI LOKAL: Semua parameter vital bergerak stabil. Tren klinis menunjukkan pemulihan."
    else:
        return "ℹ️ AI LOKAL: Fluktuasi tanda vital dalam batas toleransi normal. Tidak ada anomali tren."

# =====================================================================
# CONFIG & AUTO-MIGRATION DATABASE
# =====================================================================
DB_FILE = "telemedis_iot.db"
if 'ambang_batas' not in st.session_state:
    st.session_state['ambang_batas'] = {}
if 'gemini_response' not in st.session_state:
    st.session_state['gemini_response'] = ""

def inisialisasi_tabel_user():
    koneksi = sqlite3.connect(DB_FILE)
    kursor = koneksi.cursor()
    kursor.execute("""
        CREATE TABLE IF NOT EXISTS Users (
            id_user INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT,
            password TEXT,
            role TEXT
        )
    """)
    akun_default = [
        ('admin', 'admin', 'Resepsionis'),
        ('dokter', '12345', 'Dokter'),
        ('perawat', 'perawat', 'Perawat'),
        ('ammar', 'ammar123', 'Pasien')
    ]
    for user, pwd, role in akun_default:
        kursor.execute("SELECT * FROM Users WHERE username=?", (user,))
        if not kursor.fetchone():
            kursor.execute("INSERT INTO Users (username, password, role) VALUES (?, ?, ?)", (user, pwd, role))
    koneksi.commit()
    koneksi.close()

inisialisasi_tabel_user()

def hitung_news_status(suhu, hr, spo2, id_pasien):
    limits = st.session_state['ambang_batas'].get(id_pasien, {
        "min_hr": 60, "max_hr": 100, "min_spo2": 92, "max_suhu": 38.0
    })
    score = 0
    if spo2 < limits["min_spo2"]: score += 3
    elif spo2 <= limits["min_spo2"] + 2: score += 2
    
    if hr < limits["min_hr"] or hr > limits["max_hr"] + 20: score += 3
    elif limits["max_hr"] < hr <= limits["max_hr"] + 20: score += 2
    
    if suhu > limits["max_suhu"]: score += 2
    elif suhu < 36.0: score += 1
    
    if score >= 5 or hr < limits["min_hr"]-5 or spo2 < limits["min_spo2"]-2:
        return f"Bahaya (NEWS: {score})", score
    elif 1 <= score < 5:
        return f"Peringatan (NEWS: {score})", score
    else:
        return f"Normal (NEWS: {score})", score

# =====================================================================
# FITUR VISUALISASI ANATOMI (DIGITAL TWIN)
# =====================================================================
def render_digital_twin(suhu, hr, spo2):
    # 1. Logika Otak & Suhu (Kepala)
    warna_kepala = "#E74C3C" if suhu > 38.0 else ("#3498DB" if suhu < 36.0 else "#2ECC71")
    
    # 2. Logika Detak Jantung (Dada Tengah)
    warna_jantung = "#E74C3C" if hr > 100 or hr < 60 else "#2ECC71"
    animasi_jantung = "anim-jantung-cepat" if hr > 100 else ("anim-jantung-lambat" if hr < 60 else "anim-jantung-normal")
    
    # 3. Logika Saturasi Oksigen Paru-paru (Dada Kiri & Kanan)
    warna_paru = "#9B59B6" if spo2 < 92 else ("#F1C40F" if spo2 <= 95 else "#2ECC71")
    animasi_paru = "anim-napas-sesak" if spo2 < 95 else "anim-napas-normal"

    # PERHATIAN: Teks HTML di bawah ini sengaja dibuat mentok rata kiri 
    # agar Streamlit menggambarnya sebagai visual, bukan sebagai teks kode.
    html_twin = f"""
<style>
.twin-box {{ display: flex; justify-content: center; align-items: center; background: radial-gradient(circle, #1a252c 0%, #000000 100%); padding: 10px; border-radius: 15px; border: 1px solid #34495e; box-shadow: 0 0 15px rgba(44, 62, 80, 0.5) inset; }}
.anim-jantung-normal {{ animation: detak 1s infinite; transform-origin: center; }}
.anim-jantung-cepat {{ animation: detak 0.3s infinite; transform-origin: center; }}
.anim-jantung-lambat {{ animation: detak 2s infinite; transform-origin: center; }}
.anim-napas-normal {{ animation: napas 3s infinite; transform-origin: center; }}
.anim-napas-sesak {{ animation: napas 1s infinite; transform-origin: center; }}
@keyframes detak {{ 0% {{ transform: scale(1); }} 50% {{ transform: scale(1.3); }} 100% {{ transform: scale(1); }} }}
@keyframes napas {{ 0% {{ transform: scale(1); opacity: 0.7; }} 50% {{ transform: scale(1.05); opacity: 1; }} 100% {{ transform: scale(1); opacity: 0.7; }} }}
</style>
<div class="twin-box">
<svg width="220" height="350" viewBox="0 0 200 400" xmlns="http://www.w3.org/2000/svg">
<rect x="60" y="110" width="80" height="140" rx="30" fill="#2c3e50" opacity="0.4"/>
<rect x="75" y="230" width="20" height="110" rx="10" fill="#2c3e50" opacity="0.4"/>
<rect x="105" y="230" width="20" height="110" rx="10" fill="#2c3e50" opacity="0.4"/>
<rect x="30" y="120" width="20" height="100" rx="10" fill="#2c3e50" opacity="0.4"/>
<rect x="150" y="120" width="20" height="100" rx="10" fill="#2c3e50" opacity="0.4"/>
<circle cx="100" cy="55" r="35" fill="{warna_kepala}" opacity="0.85" filter="drop-shadow(0 0 8px {warna_kepala})"/>
<text x="100" y="60" font-family="Arial" font-size="14" fill="#000" text-anchor="middle" font-weight="bold">SUHU</text>
<ellipse cx="75" cy="145" rx="18" ry="30" fill="{warna_paru}" class="{animasi_paru}" filter="drop-shadow(0 0 5px {warna_paru})"/>
<ellipse cx="125" cy="145" rx="18" ry="30" fill="{warna_paru}" class="{animasi_paru}" filter="drop-shadow(0 0 5px {warna_paru})"/>
<text x="100" y="195" font-family="Arial" font-size="12" fill="#FFF" text-anchor="middle" opacity="0.8">O2</text>
<circle cx="100" cy="150" r="14" fill="{warna_jantung}" class="{animasi_jantung}" filter="drop-shadow(0 0 10px {warna_jantung})"/>
<text x="100" y="154" font-family="Arial" font-size="11" fill="#FFF" text-anchor="middle" font-weight="bold">HR</text>
</svg>
</div>
"""
    return html_twin
# =====================================================================
# INTEGRASI BACKEND: SERVER API FLASK 
# =====================================================================
flask_app = Flask(__name__)

@flask_app.route('/api/vital', methods=['POST'])
def terima_data_sensor():
    try:
        data = request.get_json()
        id_pasien = data.get('id_pasien')
        suhu = data.get('suhu_tubuh')
        hr = data.get('detak_jantung')
        spo2 = data.get('spo2')
        
        status = "Bahaya" if (hr > 100 or hr < 60 or spo2 < 92 or suhu > 38.0) else ("Peringatan" if (92 <= spo2 <= 95 or 37.5 < suhu <= 38.0) else "Normal")
        
        koneksi = sqlite3.connect(DB_FILE)
        kursor = koneksi.cursor()
        kursor.execute("""
            INSERT INTO TandaVital (id_pasien, suhu_tubuh, detak_jantung, spo2, status_klinis)
            VALUES (?, ?, ?, ?, ?)
        """, (id_pasien, suhu, hr, spo2, status))
        koneksi.commit()
        koneksi.close()
        return jsonify({"pesan": "Sukses", "status_klinis": status}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400

def jalankan_server_api():
    flask_app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)

if 'server_aktif' not in st.session_state:
    threading.Thread(target=jalankan_server_api, daemon=True).start()
    st.session_state['server_aktif'] = True

# =====================================================================
# FRONTEND WEBPAGE: MULTI-ROLE LOGIN INTERCEPTOR
# =====================================================================
st.set_page_config(page_title="Telemetri Medis IoT", layout="wide")

if 'logged_in' not in st.session_state:
    st.session_state['logged_in'] = False
    st.session_state['user_role'] = None
    st.session_state['username'] = None

if not st.session_state['logged_in']:
    st.markdown("<h1 style='text-align: center; color: #1B365D;'>🏥 Portal Login SIMRS Telemedis</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; color: gray;'>Sistem Informasi Manajemen Rumah Sakit & IoT Terpadu</p>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        with st.form("form_login"):
            input_user = st.text_input("Username / ID Pengguna:")
            input_pass = st.text_input("Kata Sandi (Password):", type="password")
            if st.form_submit_button("Masuk ke Sistem", use_container_width=True):
                koneksi = sqlite3.connect(DB_FILE)
                kursor = koneksi.cursor()
                kursor.execute("SELECT role FROM Users WHERE username=? AND password=?", (input_user.strip().lower(), input_pass))
                hasil = kursor.fetchone()
                koneksi.close()
                
                if hasil:
                    st.session_state['logged_in'] = True
                    st.session_state['user_role'] = hasil[0]
                    st.session_state['username'] = input_user.strip().lower()
                    st.success("Login Berhasil! Memuat dashboard...")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error("Kombinasi Username & Password tidak terdaftar!")
        
        st.info("💡 **Panduan Akun Demo:**\n* Admin: `admin` | `admin`\n* Dokter: `dokter` | `12345`\n* Pasien: `ammar` | `ammar123`")
    st.stop() 

# =====================================================================
# DASHBOARD UTAMA (SETELAH LOGIN BERHASIL)
# =====================================================================
st.title("🏥 Sistem Jaringan & Dashboard Telemetri Medis IoT")
st.markdown(f"Akses Terverifikasi: **{st.session_state['user_role']}** | Akun: `@{st.session_state['username']}`")
st.divider()

# Sistem Keamanan Menu (HIPAA & RBAC)
role_user = st.session_state['user_role']
pilihan_menu = []

if role_user in ["Dokter", "Perawat"]:
    pilihan_menu = [
        "📝 Pendaftaran Pasien Baru",
        "📊 Dashboard Real-Time & AI", 
        "🖥️ Central Monitor (Stasiun Perawat)", 
        "🗑️ Manajemen Data Pasien"
    ]
elif role_user == "Resepsionis":
    pilihan_menu = ["📝 Pendaftaran Pasien Baru"]
elif role_user == "Pasien":
    pilihan_menu = ["📊 Dashboard Real-Time & AI"]

menu = st.sidebar.selectbox("Pilih Menu Jendela:", pilihan_menu)
st.sidebar.markdown("---")

if st.sidebar.button("🚪 Keluar Sistem (Logout)", use_container_width=True, type="primary"):
    st.session_state['logged_in'] = False
    st.session_state['user_role'] = None
    st.session_state['username'] = None
    st.rerun()

st.sidebar.markdown("---")
st.sidebar.subheader("🎛️ Kontrol Sensor Perangkat")
simulasi_aktif = st.sidebar.toggle("▶️ Mulai Aliran Data Otomatis")
if simulasi_aktif:
    st.sidebar.success("Sinyal Sensor: AKTIF MENGIRIM DATA")
else:
    st.sidebar.error("Sinyal Sensor: TERPUTUS / MATI")

# ================= MENU 1: PENDAFTARAN PASIEN =================
if menu == "📝 Pendaftaran Pasien Baru":
    st.subheader("📋 Form Registrasi Pasien Baru")
    
    with st.form("form_pasien"):
        nama = st.text_input("Nama Lengkap Pasien:")
        usia = st.number_input("Usia Pasien (Tahun):", min_value=0, max_value=120, value=0)
        catatan = st.text_area("Catatan Medis / Riwayat Penyakit:")
        tombol_submit = st.form_submit_button("Daftarkan Pasien")
        
        if tombol_submit:
            if nama == "" or usia <= 0:
                st.error("Mohon isi nama dan usia dengan benar (angka lebih dari 0)!")
            else:
                username_baru = nama.split()[0].lower()
                password_baru = username_baru + "123" 
                
                koneksi = sqlite3.connect(DB_FILE)
                kursor = koneksi.cursor()
                kursor.execute("INSERT INTO Pasien (nama, usia, catatan_medis) VALUES (?, ?, ?)", (nama, usia, catatan))
                
                kursor.execute("SELECT * FROM Users WHERE username=?", (username_baru,))
                if not kursor.fetchone():
                    kursor.execute("INSERT INTO Users (username, password, role) VALUES (?, ?, ?)", (username_baru, password_baru, 'Pasien'))
                    info_akun = f"**Username:** `{username_baru}` | **Password:** `{password_baru}`"
                else:
                    username_unik = f"{username_baru}{random.randint(10,99)}"
                    password_unik = f"{username_unik}123"
                    kursor.execute("INSERT INTO Users (username, password, role) VALUES (?, ?, ?)", (username_unik, password_unik, 'Pasien'))
                    info_akun = f"**Username:** `{username_unik}` | **Password:** `{password_unik}`"
                
                koneksi.commit()
                koneksi.close()
                
                st.success(f"Pasien '{nama}' BERHASIL didaftarkan!")
                st.info(f"🔑 **AKUN LOGIN PASIEN TERCIPTA OTOMATIS!**\nBerikan data ini kepada pasien agar mereka bisa login:\n{info_akun}")

# ================= MENU 2: DASHBOARD REAL-TIME & AI =================
elif menu == "📊 Dashboard Real-Time & AI":
    st.header("🖥️ Live Monitor Tanda Vital Pasien")
    
    koneksi = sqlite3.connect(DB_FILE)
    if role_user == "Pasien":
        query_pasien = f"SELECT id_pasien, nama, usia FROM Pasien WHERE nama LIKE '%{st.session_state['username']}%' LIMIT 1"
    else:
        query_pasien = "SELECT id_pasien, nama, usia FROM Pasien"
        
    df_pasien = pd.read_sql_query(query_pasien, koneksi)
    koneksi.close()
    
    if df_pasien.empty:
        if role_user == "Pasien":
            st.warning("Akun terdeteksi, namun profil medis belum diregistrasi oleh admin.")
        else:
            st.warning("Belum ada pasien terdaftar. Silakan ke menu Pendaftaran Pasien Baru.")
    else:
        pilihan_pasien = st.selectbox("Pilih Pasien yang Akan Dipantau:", df_pasien['nama'].tolist(), disabled=(role_user == "Pasien"))
        id_terpilih = df_pasien[df_pasien['nama'] == pilihan_pasien]['id_pasien'].values[0]
        
        # Fitur Atur Ambang Batas NEWS (Khusus Tenaga Medis)
        if role_user in ["Dokter", "Perawat"]:
            with st.expander("⚙️ Kustomisasi Nilai Ambang Batas Medis Pasien (Fitur Dokter)"):
                current_limits = st.session_state['ambang_batas'].get(id_terpilih, {
                    "min_hr": 60, "max_hr": 100, "min_spo2": 92, "max_suhu": 38.0
                })
                c1, c2, c3, c4 = st.columns(4)
                new_min_hr = c1.number_input("Batas Min HR", value=current_limits["min_hr"])
                new_max_hr = c2.number_input("Batas Max HR", value=current_limits["max_hr"])
                new_min_spo2 = c3.number_input("Batas Min SpO2", value=current_limits["min_spo2"])
                new_max_suhu = c4.number_input("Batas Max Suhu", value=current_limits["max_suhu"])
                
                if st.button("💾 Simpan Batasan Khusus"):
                    st.session_state['ambang_batas'][id_terpilih] = {
                        "min_hr": new_min_hr, "max_hr": new_max_hr, "min_spo2": new_min_spo2, "max_suhu": new_max_suhu
                    }
                    st.success("Ambang batas dinamis diaktifkan!")
        
        if simulasi_aktif:
            payload = {
                "id_pasien": int(id_terpilih),
                "suhu_tubuh": round(random.uniform(36.0, 39.5), 1),
                "detak_jantung": random.randint(50, 115),
                "spo2": random.randint(88, 100)
            }
            try: requests.post("http://127.0.0.1:5000/api/vital", json=payload)
            except: pass 
        
        koneksi = sqlite3.connect(DB_FILE)
        query_medis = f"SELECT v.waktu_rekam, v.suhu_tubuh, v.detak_jantung, v.spo2, v.status_klinis, p.usia, p.catatan_medis FROM TandaVital v JOIN Pasien p ON v.id_pasien = p.id_pasien WHERE v.id_pasien = {id_terpilih} ORDER BY v.waktu_rekam DESC"
        df_medis = pd.read_sql_query(query_medis, koneksi)
        koneksi.close()
        
        if df_medis.empty:
            st.info("Menunggu data dari perangkat sensor IoT...")
        else:
            catatan_pasien = df_medis['catatan_medis'].iloc[0]
            st.write(f"**Usia:** {df_medis['usia'].iloc[0]} Tahun | **Riwayat:** {catatan_pasien}")
            
            data_terakhir = df_medis.iloc[0]
            status_news, skor_news = hitung_news_status(data_terakhir['suhu_tubuh'], data_terakhir['detak_jantung'], data_terakhir['spo2'], id_terpilih)
            
            # --- PEMBAGIAN LAYAR (KIRI: DATA & AI | KANAN: DIGITAL TWIN) ---
            kolom_kiri, kolom_kanan = st.columns([2, 1]) # Porsi layar 2:1
            
            with kolom_kiri:
                # Menampilkan Angka Vital
                c1, c2, c3 = st.columns(3)
                c1.metric("🌡️ Suhu Tubuh", f"{data_terakhir['suhu_tubuh']} °C")
                c2.metric("❤️ Detak Jantung", f"{data_terakhir['detak_jantung']} BPM")
                c3.metric("💨 Saturasi Oksigen (SpO2)", f"{data_terakhir['spo2']} %")
                
                # Menampilkan Status & Peringatan
                if "Bahaya" in status_news:
                    st.error(f"🚨 STATUS KLINIS: {status_news}")
                elif "Peringatan" in status_news:
                    st.warning(f"⚠️ STATUS KLINIS: {status_news}")
                else:
                    st.success(f"✅ STATUS KLINIS: {status_news}")
                
                # --- TAMPILAN AI LOKAL & GEMINI ---
                pesan_ai_lokal = prediksi_ai_tren(df_medis)
                if "🚨" in pesan_ai_lokal: st.error(pesan_ai_lokal)
                elif "⚠️" in pesan_ai_lokal: st.warning(pesan_ai_lokal)
                else: st.success(pesan_ai_lokal)
                
                if st.button("🔍 Minta Analisis Klinis Gemini AI", type="primary"):
                    with st.spinner("Gemini sedang menganalisis data..."):
                        hasil = minta_diagnosis_ai_gemini(data_terakhir['suhu_tubuh'], data_terakhir['detak_jantung'], data_terakhir['spo2'], catatan_pasien, pilihan_pasien)
                        st.session_state['gemini_response'] = hasil
                if st.session_state['gemini_response'] != "":
                    st.info(f"**📝 Laporan Gemini AI:**\n\n{st.session_state['gemini_response']}")

            with kolom_kanan:
                st.markdown("<h4 style='text-align: center; color: lightgray;'>🧍‍♂️ Anatomi Digital Pasien</h4>", unsafe_allow_html=True)
                
                # Memanggil mesin SVG Digital Twin dan me-render HTML-nya
                twin_html = render_digital_twin(data_terakhir['suhu_tubuh'], data_terakhir['detak_jantung'], data_terakhir['spo2'])
                st.markdown(twin_html, unsafe_allow_html=True)
                
                st.caption("🟢 Normal | 🟡 Waspada | 🔴/🟣 Kritis (Hipoksia)")
            
            # --- GRAFIK & TABEL ---
            st.markdown("### 📈 Grafik Tren Tanda Vital (Waktu ke Waktu)")
            df_grafik = df_medis.head(20).set_index('waktu_rekam')[['suhu_tubuh', 'detak_jantung', 'spo2']]
            st.line_chart(df_grafik)
            
            st.markdown("### 📋 Tabel Riwayat Rekam Medis")
            st.dataframe(df_medis, width='stretch')
            
            # --- FITUR CETAK EXCEL (DIKEMBALIKAN UTUH) ---
            if st.button("📥 Cetak Laporan Excel Premium"):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Data Rekam Medis"
                headers = ["Nama Pasien", "Usia", "Catatan Medis", "Waktu Rekam", "Suhu (°C)", "Detak Jantung (BPM)", "SpO2 (%)", "Status Klinis"]
                
                fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
                font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                
                for c_num, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=c_num, value=h)
                    cell.fill = fill_header; cell.font = font_header; cell.alignment = Alignment(horizontal="center")
                
                lebar_kolom = {'A': 25, 'B': 10, 'C': 35, 'D': 25, 'E': 15, 'F': 22, 'G': 15, 'H': 20}
                for kolom, lebar in lebar_kolom.items():
                    ws.column_dimensions[kolom].width = lebar
                
                koneksi = sqlite3.connect(DB_FILE)
                kursor = koneksi.cursor()
                kursor.execute(f"SELECT p.nama, p.usia, p.catatan_medis, v.waktu_rekam, v.suhu_tubuh, v.detak_jantung, v.spo2, v.status_klinis FROM TandaVital v JOIN Pasien p ON v.id_pasien = p.id_pasien WHERE v.id_pasien = {id_terpilih}")
                for r_idx, row in enumerate(kursor.fetchall(), 2):
                    for c_idx, val in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=val).alignment = Alignment(horizontal="center")
                koneksi.close()
                
                wb.save("Laporan_Web_Telemedis.xlsx")
                st.success("Sukses! File 'Laporan_Web_Telemedis.xlsx' berhasil dibuat di folder projek Anda dengan rapi.")
                
        if simulasi_aktif:
            time.sleep(2.5) 
            st.rerun()      
        else:
            if st.button("🔄 Refresh Data Manual"):
                st.rerun()  

# ================= MENU 3: CENTRAL MONITOR (STASIUN PERAWAT) =================
elif menu == "🖥️ Central Monitor (Stasiun Perawat)":
    st.header("🖥️ Stasiun Monitor Sentral ICU (Multi-Bed Grid View)")
    st.markdown("*(Memantau kondisi seluruh bangsal/kamar pasien secara simultan dari satu layar komputer)*")
    
    koneksi = sqlite3.connect(DB_FILE)
    df_pasien = pd.read_sql_query("SELECT id_pasien, nama, usia FROM Pasien", koneksi)
    
    if df_pasien.empty:
        st.warning("Belum ada pasien terdaftar di dalam database.")
    else:
        if simulasi_aktif:
            for index, pasien in df_pasien.iterrows():
                try:
                    requests.post("http://127.0.0.1:5000/api/vital", json={
                        "id_pasien": int(pasien['id_pasien']),
                        "suhu_tubuh": round(random.uniform(36.0, 39.5), 1),
                        "detak_jantung": random.randint(55, 115),
                        "spo2": random.randint(89, 100)
                    })
                except: pass
        
        kolom_grid = st.columns(3)
        for idx, pasien in df_pasien.iterrows():
            df_vital = pd.read_sql_query(f"SELECT suhu_tubuh, detak_jantung, spo2 FROM TandaVital WHERE id_pasien={pasien['id_pasien']} ORDER BY waktu_rekam DESC LIMIT 1", koneksi)
            
            with kolom_grid[idx % 3]:
                if not df_vital.empty:
                    v_data = df_vital.iloc[0]
                    st_news, sk_news = hitung_news_status(v_data['suhu_tubuh'], v_data['detak_jantung'], v_data['spo2'], pasien['id_pasien'])
                    color_box = "#D98880" if "Bahaya" in st_news else ("#F4D03F" if "Peringatan" in st_news else "#52BE80")
                    text_color = "#E74C3C" if "Bahaya" in st_news else ("#F39C12" if "Peringatan" in st_news else "#27AE60")
                    
                    st.markdown(f"""
                        <div style='border: 2px solid {color_box}; padding: 15px; border-radius: 8px; margin-bottom: 15px; background-color: #1A1A1A;'>
                            <h4 style='margin:0; color:{text_color};'>{pasien['nama']} (BED {pasien['id_pasien']})</h4>
                            <p style='margin:0; font-size:13px; color:#BDC3C7;'>Usia: {pasien['usia']} Thn | Status: <b>{st_news}</b></p>
                            <hr style='margin:8px 0; border-color:#444;'>
                            <table style='width:100%; text-align:center; color:white; font-size:14px;'>
                                <tr>
                                    <td>❤️ {v_data['detak_jantung']} <br><span style='font-size:10px; color:#888;'>BPM</span></td>
                                    <td>💨 {v_data['spo2']}% <br><span style='font-size:10px; color:#888;'>SpO2</span></td>
                                    <td>🌡️ {v_data['suhu_tubuh']}°C <br><span style='font-size:10px; color:#888;'>Suhu</span></td>
                                </tr>
                            </table>
                        </div>
                    """, unsafe_allow_html=True)
                else:
                    st.info(f"🛏️ Bed {pasien['id_pasien']}: {pasien['nama']} (Menunggu Sinyal...)")
                    
    koneksi.close()
    if simulasi_aktif:
        time.sleep(2.5)
        st.rerun()

# ================= MENU 4: MANAJEMEN DATA =================
elif menu == "🗑️ Manajemen Data Pasien":
    st.subheader("🗑️ Hapus Data Pasien & Rekam Medis")
    st.warning("PERHATIAN: Menghapus data pasien di bawah ini juga akan menghapus seluruh riwayat grafik dan tabel rekam medis pasien tersebut secara permanen!")
    
    koneksi = sqlite3.connect(DB_FILE)
    df_pasien = pd.read_sql_query("SELECT id_pasien, nama FROM Pasien", koneksi)
    koneksi.close()
    
    if df_pasien.empty:
        st.info("Sistem bersih. Belum ada pasien yang terdaftar di database.")
    else:
        pilihan_hapus = st.selectbox("Pilih Nama Pasien yang Ingin Dihapus:", df_pasien['nama'].tolist())
        id_hapus = df_pasien[df_pasien['nama'] == pilihan_hapus]['id_pasien'].values[0]
        
        if st.button(f"🚨 Hapus Permanen Pasien '{pilihan_hapus}'", type="primary"):
            koneksi = sqlite3.connect(DB_FILE)
            kursor = koneksi.cursor()
            kursor.execute(f"DELETE FROM TandaVital WHERE id_pasien = {id_hapus}")
            kursor.execute(f"DELETE FROM Pasien WHERE id_pasien = {id_hapus}")
            koneksi.commit()
            koneksi.close()
            
            st.success(f"Berhasil! Pasien '{pilihan_hapus}' dan seluruh data medisnya telah dihapus dari sistem.")
            time.sleep(2) 
            st.rerun()